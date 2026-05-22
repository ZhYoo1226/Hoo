import json
import os
import re
from datetime import datetime

import openpyxl
import pandas as pd

from .global_functions import GlobalFunction


class ExcelEntity:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    def read_sheets(self, line: int = -1):
        # 打开Excel文件
        workbook = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            rows_data = []
            ws = workbook[sheet_name]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if line != -1 and i >= line:
                    break
                rows_data.append(row)
            yield sheet_name, rows_data
        workbook.close()

    def write_sheet_markdown(self, sheet_name: str, rows_data: list, output_path: str):
        """
        将sheet数据写入markdown表格文件
        """
        # 收集所有内容到列表
        content = []
        content.append(f"### {sheet_name}\n\n")
        if not rows_data:
            content.append("此sheet为空\n")
            # 一次性写入
            with open(output_path, 'w', encoding='utf-8') as f:
                f.writelines(content)
            return content
        # 创建markdown表格
        header = rows_data[0]
        # 写入表头
        content.append('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in header) + ' |\n')
        # 写入分隔线
        content.append('| ' + ' | '.join(['---'] * len(header)) + ' |\n')
        # 写入数据行
        for row in rows_data[1:]:
            content.append('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in row) + ' |\n')
        content.append('\n')
        # 一次性写入文件
        with open(output_path, 'w', encoding='utf-8') as f:
            f.writelines(content)
        return content


# FIXME mim：信息从哪里来，组织文件状态。
class WorkFile:
    """用户发送来的工作文件，对应 files_abstract.parquet 中的一行记录"""

    """
    工作文件是否绑定任务，还是任务绑定工作文件。
    文件在对话中的显示状态,to_markdown()
    文件对file_abstract的属性读取，比如是否完成分析。
    """
    def __init__(self, file_abstract: dict):
        #命名的一致性
        self.uuid = file_abstract.get("uuid", "")
        self.file_path = file_abstract.get("文件路径", "")
        self.file_name = file_abstract.get("文件名称", "")
        self.hash_value = file_abstract.get("hash值", "")
        self.updated_at = file_abstract.get("更新时间", "")
        self.preview = file_abstract.get("文件预览", "")
        self.abstract = file_abstract.get("文件摘要", "")
        self.entities = file_abstract.get("关键实体", "")
        self.importance = file_abstract.get("重要性", "")
        self.file_abstract = file_abstract

    def __getitem__(self, item):
        return self.file_abstract.get(item, None)

    @property
    def abstract_filled(self) -> bool:
        return bool(self.abstract and self.abstract.strip())

    def to_dict(self) -> dict:
        return self.file_abstract.copy()

    def to_markdown(self) -> str:
        """文件在对话中的显示状态"""
        status = "已摘要" if self.abstract_filled else "待摘要"
        return f" - 文件：{self['文件名称']}(uuid:{self['文件路径']})"

    def __repr__(self):
        return f"WorkFile(uuid={self.uuid}, name={self.file_name})"


class WorkTask:
    def __init__(self, json: dict, owner):
        self.task_json = json
        self.owner = owner
        # 动作列表为空
        self._action_list = []

    def task(self):
        return self.task_json['task']

    @property
    def task_id(self):
        return self.task_json['task_id']

    @property
    def goal(self):
        return self.task_json['goal']

    @property
    def type(self):
        return self.task_json['type']

    @property
    def name(self):
        return self.task_json['name']

    @property
    def finished(self):
        return bool(self.task_json.get('finished', False))

    def __json__(self):
        return self.task_json

    def __str__(self):
        return json.dumps(self.task_json, ensure_ascii=False, indent=2)

    def __repr__(self):
        return self.__str__()

    def __eq__(self, other):
        return self.task_json['task'] == other.task_json['task']

    @property
    def action_list(self):
        return self._action_list

    def to_markdown(self):
        """将任务信息转换为markdown格式"""
        return f" - 任务: {self.name}({self.task_id})"

    def task_execute_history(self):
        """任务执行历史记录"""
        executor_path = GlobalFunction.task_id_to_executor_path(self.task_id)
        # 先读取文件内容，使用r模式，如果文件不存在，返回None
        file_bytes = "任务执行历史记录可能存在问题，需要告知用户，或者取消掉本任务。"
        if os.path.exists(executor_path):
            with open(executor_path, 'r', encoding='utf-8') as f:
                file_bytes = f.read()
        return file_bytes

    def update_task_record(self):
        """更新任务记录到任务执行.md文件"""
        executor_path = GlobalFunction.task_id_to_executor_path(self.task_id)
        os.makedirs(os.path.dirname(executor_path), exist_ok=True)
        # 先读取文件内容，使用r模式
        file_bytes = None
        if os.path.exists(executor_path):
            with open(executor_path, 'r', encoding='utf-8') as f:
                file_bytes = f.read()

        if file_bytes:
            # 找到内容中第一个“# 任务执行”
            start_index = file_bytes.find("# 任务执行")
            if start_index != -1:
                append_bytes = file_bytes[start_index:]
            else:
                # 如果没有找到"# 任务执行"，保留原有全部内容
                append_bytes = file_bytes
        else:
            # 没有内容，说明是新建
            append_bytes = ["# 任务执行"]
            hist_text = self.owner.get_chat_history_text(limit=5, roles=['用户', '助手', '行动', '观察'])
            append_bytes.append(hist_text)

        # 重新以w模式打开写入新内容
        with open(executor_path, 'w', encoding='utf-8') as f:
            # 写入新的内容
            f.write(f"# 任务信息摘要\n\n")
            f.write(f"```json\n{json.dumps(self.task_json, ensure_ascii=False, indent=2)}\n```\n\n")
            f.write("\n".join(append_bytes))

    @action_list.setter
    def action_list(self, value):
        self._action_list = value


class ChatMessage:
    """聊天消息类
    包含role、text、timestamp三个字段
    """
    # 提前定义映射关系，避免每次初始化都创建字典
    _role_map = {
        "用户": "user",
        "助手": "assistant",
        "系统": "system",
        "行动": "action",
        "观察": "observation",
        "思考": "think"
    }

    def __init__(self, role: str = None, content: str = None, data: str = None):
        # 如果role为中文需要转换为英文,用户:user,助手:assistant,系统:system,行动:action,观察:observation,执行:action
        self.role = role
        self.content = content
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 当前时间（时分秒格式）
        if data:
            self.parse(data)

    def print(self):
        print(self)

    def __str__(self):
        # {self.timestamp}
        return f"[{self.role}]: {self.content}"

    def __repr__(self):
        return self.__str__()

    def to_conversation(self):
        """转换为对话格式"""
        return {"role": self._role_map[self.role] if self.role in self._role_map else self.role,
                "content": self.content}

    def parse(self, data: str):
        """读取消息文本"""
        # [self.role]: {self.text} 格式，解析data
        pattern = r'^\[(.*?)\]\:\s*(.*)$'
        match = re.match(pattern, data.strip())
        if match:
            self.role = match.group(1).strip()
            self.content = match.group(2).strip()
        else:
            # 如果不匹配格式，保留原逻辑，整个内容作为文本
            self.content = data.strip()
        return self

    def to_dict(self):
        """Convert ChatMessage instance to a dictionary for JSON serialization."""
        return {
            "role": self.role,
            "content": self.content
        }

    def __json__(self):
        """Return a JSON-serializable representation of the object."""
        return self.to_dict()

    @classmethod
    def from_dict(cls, data):
        """Create a ChatMessage instance from a dictionary."""
        return cls(role=data.get('role'), content=data.get('content'))

    def __getitem__(self, key):
        """使ChatMessage对象支持字典风格访问"""
        if key == 'role':
            return self.role
        elif key == 'content':
            return self.content
        elif key == 'timestamp':
            return self.timestamp
        else:
            raise KeyError(f"'{key}' is not a valid attribute of ChatMessage")

    def __setitem__(self, key, value):
        """使ChatMessage对象支持字典风格赋值"""
        if key == 'role':
            self.role = value
        elif key == 'content':
            self.content = value
        elif key == 'timestamp':
            self.timestamp = value
        else:
            raise KeyError(f"'{key}' is not a valid attribute of ChatMessage")


class PromptIndex:
    _by_name = {}
    _by_filename = {}

    def __init__(self, index_list):
        """
        提示词对象格式：
        {
            "名称": name,
            "文件名": filename,
            "文件路径": os.path.relpath(file_path, workdir),  # 保存相对路径便于后续查找
            "输入参数": {p: "input" for p in placeholders} if placeholders else {}
        }
        初始化索引，接收原始列表并构建两个映射字典。
        :param index_list: 包含提示词信息的列表，每个元素为字典
        """
        for item in index_list:
            name = item.get("名称")
            filename = item.get("文件名")
            if name:
                PromptIndex._by_name[name] = item
            if filename:
                PromptIndex._by_filename[filename] = item

    def to_dict(self):
        """转换为可序列化字典"""
        return {
            "_by_name": self._by_name,
            "_by_filename": self._by_filename
        }

    def __str__(self):
        """支持json序列化打印"""
        return json.dumps(self.to_dict(), ensure_ascii=False)

    def __repr__(self):
        """支持json序列化打印"""
        return self.__str__()

    def __getstate__(self):
        """支持pickle序列化"""
        return self.to_dict()

    def __setstate__(self, state):
        """支持pickle反序列化"""
        self._by_name = state.get("_by_name", {})
        self._by_filename = state.get("_by_filename", {})

    @classmethod
    def from_dict(cls, data):
        """从字典反序列化创建对象"""
        instance = cls([])
        instance._by_name = data.get("_by_name", {})
        instance._by_filename = data.get("_by_filename", {})
        return instance

    # 加载提示词，填充占位符
    def load_prompt(self, prompt_name: str, owner, state) -> tuple[str, dict]:
        prompt_json = self.get_src_prompt(prompt_name)
        if not prompt_json:
            return None, None
        input_param = prompt_json["输入参数"]
        # 使用完整路径（包含子目录相对路径）加载文件
        from common import GlobalFunction
        output_text = GlobalFunction.load_file(os.path.join(GlobalFunction.prompt_path(), prompt_json["文件路径"]))
        if input_param:
            # 遍历每个需要替换的键
            for key in input_param.keys():
                placeholder = "{{" + key + "}}"  # 占位符格式 {{key}}
                value = None
                if state:
                    # 优先从state状态属性中获取
                    value = state.read_property(key)
                # 如果状态属性中没有，再从实体属性中获取
                if value is None:
                    value = owner.read_property(key)

                if value is None:
                    value = "(Not Provided)"
                    # FIXME 需要构建提示词创建这个参数获得的代码。
                    GlobalFunction.log("运行时错误", f"提示词[{prompt_name}]输入参数[{key}]不存在，请检查提示词配置。")
                # 替换所有出现的占位符
                output_text = output_text.replace(placeholder, str(value))
        return output_text, prompt_json

    # 加载原始提示词
    def get_src_prompt(self, name) -> dict:
        """
        统一查找接口，支持名称或文件名（含路径）。
        规则：
        - 如果输入以 '.md' 结尾，视为文件名（自动提取最后文件名部分），按文件名查找。
        - 否则视为名称，按名称查找。
        :param name: 名称或文件名（可带路径）
        :return: 对应的条目字典，未找到返回 None
        """
        # 提取基本文件名（处理路径）
        filename = os.path.basename(name)
        # 判断是否为 markdown 文件（可扩展其他后缀）
        if filename.endswith('.md'):
            return self.get_by_filename(filename)
        else:
            return self.get_by_name(name)

    def get_by_name(self, name) -> dict:
        """
        根据名称查找提示词条目。
        :param name: 提示词的“名称”字段
        :return: 对应的条目字典，若不存在则返回 None
        """
        return self._by_name.get(name)

    def get_by_filename(self, filename) -> dict:
        """
        根据文件名查找提示词条目。
        :param filename: 提示词的“文件名”字段
        :return: 对应的条目字典，若不存在则返回 None
        """
        return self._by_filename.get(filename)

    def get_all(self) -> list[dict]:
        """返回所有条目列表（可选）"""
        return list(self._by_name.values())
