import ast
import json
import os
import re
import xml.etree.ElementTree as ET
from configparser import ConfigParser

import toml
from docx import Document
from ruamel.yaml import YAML
from openpyxl import load_workbook

from .base_state import BaseState
from .state_modify_support import (
    _atomic_replace,
    _backup_file,
    _rollback_file,
    _set_by_path,
    _log,
    _verify_file_type,
)


# --------------------------------------------------------------------------
# 格式化文件修改状态
# --------------------------------------------------------------------------

class ModifyYamlState(BaseState):
    """修改 YAML 文件指定路径的值。xpath 为点号分隔路径，如 database.port"""
    stateName = "ModifyYamlState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = value
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            ryaml = YAML()
            with open(self.file, "r", encoding="utf-8") as f:
                data = ryaml.load(f) or {}
            _set_by_path(data, self.xpath, self.value)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                ryaml.dump(data, f)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "yaml")
            _log(owner, "info", f"YAML 修改成功: {self.file} → {self.xpath} = {self.value}")
        except Exception as e:
            _log(owner, "error", f"YAML 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyJsonState(BaseState):
    """修改 JSON 文件指定路径的值。xpath 为点号分隔路径，如 server.port"""
    stateName = "ModifyJsonState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = value
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            with open(self.file, "r", encoding="utf-8") as f:
                data = json.load(f)
            _set_by_path(data, self.xpath, self.value)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "json")
            _log(owner, "info", f"JSON 修改成功: {self.file} → {self.xpath} = {self.value}")
        except Exception as e:
            _log(owner, "error", f"JSON 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyIniState(BaseState):
    """修改 INI/Properties 文件指定键的值。xpath 为 section.key 或 key"""
    stateName = "ModifyIniState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = str(value)
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            cp = ConfigParser()
            cp.read(self.file, encoding="utf-8")
            if "." in self.xpath:
                section, key = self.xpath.split(".", 1)
            else:
                section, key = "DEFAULT", self.xpath
            if section != "DEFAULT" and not cp.has_section(section):
                cp.add_section(section)
            cp.set(section, key, self.value)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                cp.write(f)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "ini")
            _log(owner, "info", f"INI 修改成功: {self.file} → [{section}] {key} = {self.value}")
        except Exception as e:
            _log(owner, "error", f"INI 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyTomlState(BaseState):
    """修改 TOML 文件指定路径的值。xpath 为点号分隔路径，如 database.port"""
    stateName = "ModifyTomlState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = value
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            with open(self.file, "r", encoding="utf-8") as f:
                data = toml.load(f)
            _set_by_path(data, self.xpath, self.value)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                toml.dump(data, f)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "toml")
            _log(owner, "info", f"TOML 修改成功: {self.file} → {self.xpath} = {self.value}")
        except Exception as e:
            _log(owner, "error", f"TOML 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyExcelState(BaseState):
    """修改 Excel 文件指定单元格的值。xpath 为 SheetName!CellRef 格式，如 Sheet1!B5"""
    stateName = "ModifyExcelState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = value
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            wb = load_workbook(self.file)
            if "!" in self.xpath:
                sheet_name, cell_ref = self.xpath.split("!", 1)
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表不存在: {sheet_name}")
                ws = wb[sheet_name]
            else:
                ws = wb.active
                cell_ref = self.xpath
            ws[cell_ref] = self.value
            _atomic_replace(self.file, lambda p: wb.save(p))
            _verify_file_type(self.file, "xlsx")
            _log(owner, "info", f"Excel 修改成功: {self.file} → {self.xpath} = {self.value}")
        except Exception as e:
            _log(owner, "error", f"Excel 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


# --------------------------------------------------------------------------
# 非格式化文件修改状态
# --------------------------------------------------------------------------

class ModifyHtmlState(BaseState):
    """修改 HTML 文件中的文本。origin_text → new_text 纯文本替换"""
    stateName = "ModifyHtmlState"

    def __init__(self, file: str, origin_text: str, new_text: str, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.origin_text = origin_text
        self.new_text = new_text
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            with open(self.file, "r", encoding="utf-8") as f:
                content = f.read()
            content = content.replace(self.origin_text, self.new_text)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                f.write(content)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "html")
            _log(owner, "info", f"HTML 修改成功: {self.file}")
        except Exception as e:
            _log(owner, "error", f"HTML 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyMarkdownState(BaseState):
    """修改 Markdown 文件中的文本。origin_text → new_text 纯文本替换"""
    stateName = "ModifyMarkdownState"

    def __init__(self, file: str, origin_text: str, new_text: str, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.origin_text = origin_text
        self.new_text = new_text
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            with open(self.file, "r", encoding="utf-8") as f:
                content = f.read()
            content = content.replace(self.origin_text, self.new_text)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                f.write(content)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "md")
            _log(owner, "info", f"Markdown 修改成功: {self.file}")
        except Exception as e:
            _log(owner, "error", f"Markdown 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyDocxState(BaseState):
    """修改 DOCX 文件中的文本。在 paragraph run 层级做 origin_text → new_text 替换"""
    stateName = "ModifyDocxState"

    def __init__(self, file: str, origin_text: str, new_text: str, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.origin_text = origin_text
        self.new_text = new_text
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            doc = Document(self.file)
            found = False
            for p in doc.paragraphs:
                for run in p.runs:
                    if self.origin_text in run.text:
                        run.text = run.text.replace(self.origin_text, self.new_text)
                        found = True
            if not found:
                raise ValueError(f"未找到匹配文本: {self.origin_text}")
            _atomic_replace(self.file, lambda p: doc.save(p))
            _verify_file_type(self.file, "docx")
            _log(owner, "info", f"DOCX 修改成功: {self.file}")
        except Exception as e:
            _log(owner, "error", f"DOCX 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


# --------------------------------------------------------------------------
# XML / .env / 正则 / 代码文件修改状态
# --------------------------------------------------------------------------

def _set_xml_by_path(element, path: str, value):
    """在 XML 元素树中按点号路径设值。末尾 @attr 设属性，否则设子元素文本。"""
    keys = path.split(".")
    target = element
    for key in keys[:-1]:
        child = target.find(key)
        if child is None:
            child = ET.SubElement(target, key)
        target = child
    last = keys[-1]
    if last.startswith("@"):
        target.set(last[1:], str(value))
    else:
        target.text = str(value)


class ModifyXmlState(BaseState):
    """修改 XML 文件。xpath 为点号路径，末尾 @attr 设属性，否则设文本"""
    stateName = "ModifyXmlState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = value
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            tree = ET.parse(self.file)
            _set_xml_by_path(tree.getroot(), self.xpath, self.value)
            tmp = self.file + ".tmp"
            tree.write(tmp, encoding="utf-8", xml_declaration=True)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "xml")
            _log(owner, "info", f"XML 修改成功: {self.file} → {self.xpath} = {self.value}")
        except Exception as e:
            _log(owner, "error", f"XML 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyEnvState(BaseState):
    """修改 .env 文件指定键的值。xpath 为键名，不存在则追加"""
    stateName = "ModifyEnvState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = str(value)
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            with open(self.file, "r", encoding="utf-8") as f:
                lines = f.readlines()
            found = False
            for i, line in enumerate(lines):
                stripped = line.strip()
                if stripped.startswith("#") or stripped == "":
                    continue
                if "=" in stripped:
                    k, _ = stripped.split("=", 1)
                    if k.strip() == self.xpath:
                        lines[i] = f"{self.xpath}={self.value}\n"
                        found = True
                        break
            if not found:
                lines.append(f"\n{self.xpath}={self.value}\n")
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                f.writelines(lines)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "env")
            _log(owner, "info", f".env 修改成功: {self.file} → {self.xpath} = {self.value}")
        except Exception as e:
            _log(owner, "error", f".env 修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyRegexState(BaseState):
    """正则替换文件内容。origin_text 为正则模式，new_text 为替换文本"""
    stateName = "ModifyRegexState"

    def __init__(self, file: str, origin_text: str, new_text: str, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.origin_text = origin_text
        self.new_text = new_text
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            with open(self.file, "r", encoding="utf-8") as f:
                content = f.read()
            content = re.sub(self.origin_text, self.new_text, content)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                f.write(content)
            os.replace(tmp, self.file)
            _log(owner, "info", f"正则替换成功: {self.file}")
        except Exception as e:
            _log(owner, "error", f"正则替换失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass


class ModifyPyCodeState(BaseState):
    """修改 Python 源文件中的赋值语句。xpath 为变量名，value 为新值。"""
    stateName = "ModifyPyCodeState"

    def __init__(self, file: str, xpath: str, value, **kwargs):
        super().__init__(**kwargs)
        self.file = file
        self.xpath = xpath
        self.value = value
        self.backup_path = None

    def Enter(self, owner):
        owner._uuid_maps[self._uuid] = self.file

    def Execute(self, owner):
        self.backup_path = _backup_file(self.file)
        try:
            with open(self.file, "r", encoding="utf-8") as f:
                content = f.read()
            ast.parse(content)
            pattern = rf"^(\s*{re.escape(self.xpath)}\s*=\s*).+"
            new_content, count = re.subn(
                pattern, rf"\g<1>{repr(self.value)}", content, count=1, flags=re.MULTILINE
            )
            if count == 0:
                raise ValueError(f"未找到变量赋值: {self.xpath}")
            ast.parse(new_content)
            tmp = self.file + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                f.write(new_content)
            os.replace(tmp, self.file)
            _verify_file_type(self.file, "py")
            _log(owner, "info", f"Python 代码修改成功: {self.file} → {self.xpath} = {self.value}")
        except Exception as e:
            _log(owner, "error", f"Python 代码修改失败: {e}")
            _rollback_file(self.file, self.backup_path)
        owner.remove_state(self)

    def Exit(self, owner):
        pass
